from fastapi import FastAPI, HTTPException, UploadFile, File, Form, Query
from models import Student, Project
from openpyxl import Workbook, load_workbook
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
import shutil
import os
import openpyxl
from typing import List
import pandas as pd
import logging
import math 


app = FastAPI()

origins = [
    "http://localhost:5173",
    "http://127.0.0.1:5173",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,  
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.mount("/static", StaticFiles(directory="static"), name="static")

UPLOAD_DIR = "uploads"
EXCEL_FILE = "projects.xlsx"
os.makedirs(UPLOAD_DIR, exist_ok=True)

def load_students():
    if not os.path.exists('students.xlsx'):
        raise HTTPException(status_code=404, detail="Файл студентов не найден")
    students_df = pd.read_excel('students.xlsx')
    return students_df


@app.get("/scratch/{filename}")
def get_scratch(filename: str):
    file_path = os.path.join("static", filename)
    
    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Файл не найден")
    
    return FileResponse(file_path)

def init_excel_file(file_name: str, headers: list):
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        wb.save(file_name)

init_excel_file("students.xlsx", ["ID", "ФИО", "Класс", "Учитель", "Контакты"])
init_excel_file("projects.xlsx", [
    "ID", "Название", "Описание", "ID Автора", "Файл", "Лайки", "Отзывы", "Просмотры", "Пожелания"
])

# Чтение данных из Excel
def load_students():
    students_df = pd.read_excel('students.xlsx')
    return students_df

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@app.get("/students/search")
def search_students(query: str = "", class_name: str = "", teacher_name: str = ""):
    try:
        students_df = load_students()
        
        # Применяем фильтрацию по запросу
        result = students_df[students_df["ФИО"].str.contains(query, case=False, na=False)] if query else students_df
        
        if class_name:
            result = result[result["Класс"].str.contains(class_name, case=False, na=False)]
        
        if teacher_name:
            result = result[result["Учитель"].str.contains(teacher_name, case=False, na=False)]
        
        # Обработка NaN значений перед преобразованием в список
        result = result.applymap(lambda x: '' if isinstance(x, float) and math.isnan(x) else x)
        
        # Преобразуем в список словарей для ответа
        students_list = result[["ID", "ФИО", "Класс", "Учитель", "Контакты"]].rename(columns={
            "ID": "id",
            "ФИО": "full_name",
            "Класс": "class_name",
            "Учитель": "teacher",
            "Контакты": "contacts"
        }).to_dict(orient="records")
        
        return students_list
    
    except Exception as e:
        logger.error(f"Ошибка при обработке запроса: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Ошибка на сервере: {str(e)}")

@app.post("/projects/")
async def upload_project(
    title: str = Form(...),
    description: str = Form(...),
    student_id: int = Form(...),
    file: UploadFile = File(...)
):
    # Сохраняем HTML-файл
    file_location = os.path.join(UPLOAD_DIR, file.filename)
    with open(file_location, "wb") as f:
        shutil.copyfileobj(file.file, f)

    # Открываем Excel
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    next_id = ws.max_row  # предполагаем, что первая строка — заголовки

    # Добавляем запись
    ws.append([
        next_id,        # ID
        title,          # Название
        description,    # Описание
        student_id,     # ID Автора
        file.filename,  # Имя файла
        0,              # Лайки
        "",             # Отзывы
        0,              # Просмотры
        ""              # Пожелания
    ])
    wb.save(EXCEL_FILE)

    return JSONResponse({"message": "Проект успешно сохранён", "project_id": next_id})

@app.post("/add-student/")
def add_student(student: Student):
    file_path = "students.xlsx"

    # Если файл существует — загружаем, иначе создаём новый
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["ID", "ФИО", "Класс", "Учитель", "Контакты"])  # заголовки

    # Определим ID автоматически (по количеству строк)
    new_id = ws.max_row  # предполагается, что первая строка — заголовок

    ws.append([
        new_id,
        student.full_name,
        student.class_name,
        student.teacher,
        str(student.contacts) if student.contacts else ""
    ])

    wb.save(file_path)
    return {"message": f"Ученик успешно добавлен с ID {new_id}"}

@app.post("/add-project/")
def add_project(project: Project):
    wb = load_workbook("projects.xlsx")
    ws = wb.active
    ws.append([
        project.id,
        project.name,
        project.description,
        project.author_id,
        project.likes,
        str(project.reviews),
        project.views,
        ", ".join(project.wishes)
    ])
    wb.save("projects.xlsx")
    return {"message": "Проект успешно добавлен в Excel"}